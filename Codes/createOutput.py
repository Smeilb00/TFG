import openpyxl
import openpyxl.worksheet.dimensions
import pandas as pd
import os.path

rutaOutput = '../Output/Output.xlsx'

def run(insert, nameOut):
    if nameOut != "def":
        global rutaOutput
        rutaOutput = nameOut

    if os.path.exists(rutaOutput):
        df = import_excel()
        update_excel(insert, nameOut)
    else:
        create_excel(insert)

def import_excel():
    df = pd.read_excel(rutaOutput)
    return df

def update_excel(newdata, nameOut):

    writer = pd.ExcelWriter(rutaOutput, engine='openpyxl')
    writer.book = openpyxl.load_workbook(rutaOutput)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    if nameOut == "def":
        nameOut = rutaOutput
    reader = pd.read_excel(r""+nameOut)

    newdata.to_excel(writer, sheet_name='Data', index=False, header=False, startrow=len(reader)+1)
    writer.close()

def create_excel(df):
    writer = pd.ExcelWriter(rutaOutput, engine='openpyxl')
    df.to_excel(writer, sheet_name='Data', index=False)

    writer.save()

    other_sheet()

def other_sheet():
    wb = openpyxl.load_workbook(rutaOutput)
    writer = pd.ExcelWriter(rutaOutput, engine='openpyxl')

    data = wb['Data']
    data.column_dimensions['A'].width = 20
    data.column_dimensions['B'].width = 20
    data.column_dimensions['C'].width = 30

    calc = wb.create_sheet('Calculos')
    calc.column_dimensions['B'].width = 25
    calc.column_dimensions['C'].width = 25
    calc.column_dimensions['D'].width = 25
    calc.title = "Calculos"

    wb.active = wb["Calculos"]

    calc['B10'] = 'Algoritmos'
    calc['B11'] = 'linear'
    calc['B12'] = 'rbf'
    calc['B13'] = 'poly'
    calc['B14'] = 'Face_recognition'

    calc['C10'] = 'Tiempo medio'
    calc['C11'] = '=AVERAGEIF(Data!A:A,B11,Data!C:C)'
    calc['C12'] = '=AVERAGEIF(Data!A:A,B12,Data!C:C)'
    calc['C13'] = '=AVERAGEIF(Data!A:A,B13,Data!C:C)'
    calc['C14'] = '=AVERAGEIF(Data!A:A,B14,Data!C:C)'

    calc['D10'] = 'Tasa de acierto'
    calc['D11'] = '=((COUNTIFS(Data!A:A,B11,Data!B:B,"si"))/((COUNTIF(Data!A:A,B11))-(COUNTIFS(Data!B:B,"unknown",Data!A:A,B11))))'
    calc['D11'].number_format = '0%'
    calc['D12'] = '=((COUNTIFS(Data!A:A,B12,Data!B:B,"si"))/((COUNTIF(Data!A:A,B12))-(COUNTIFS(Data!B:B,"unknown",Data!A:A,B12))))'
    calc['D12'].number_format = '0%'
    calc['D13'] = '=((COUNTIFS(Data!A:A,B13,Data!B:B,"si"))/((COUNTIF(Data!A:A,B13))-(COUNTIFS(Data!B:B,"unknown",Data!A:A,B13))))'
    calc['D13'].number_format = '0%'
    calc['D14'] = '=((COUNTIFS(Data!A:A,B14,Data!B:B,"si"))/((COUNTIF(Data!A:A,B14))-(COUNTIFS(Data!B:B,"unknown",Data!A:A,B14))))'
    calc['D14'].number_format = '0%'

    wb.save(rutaOutput)
    wb.close()